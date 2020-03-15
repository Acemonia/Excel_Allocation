import math
import os
from openpyxl import load_workbook

import tkinter as tk
from tkinter import filedialog

# 初步分配
def allocate(sheet_area, total, ratio):
    for row in sheet_area:
        for cell in row:
            cell.value /= times
            if math.modf(cell.value)[0] > ratio:
                cell.value = math.ceil(cell.value)  # 向上取整
            else:
                cell.value = int(cell.value)  # 向下取整
                if ratio > 0.5 and cell.value == 0:
                    cell.value = 1
            total = total + cell.value
    return total

# 调整成固定值
def allo_counts(sht, total, counts, half, hongyan_flag):
    if not hongyan_flag:
        start_pos = sht['D3']
        sht_area = sht['D3':'M22']
        sht_idle_area = sht['D12':'M13']
        sht_female_idle_area = sht['D12':'M22']
        sht_male_idle_area = sht['D3':'M13']
    else:
        start_pos = sht['E3']
        sht_area = sht['E3':'N22']
        sht_idle_area = sht['E12':'N13']
        sht_female_idle_area = sht['E12':'E22']
        sht_male_idle_area = sht['E3':'E13']

    if total > counts :
        max = start_pos.value
        for row in sht_area:
            if row in sht_idle_area:
                continue
            for cell in row:
                if max < cell.value:
                    max = cell.value

        exit_flag = False
        while total > counts:
            for row in sht_area:
                if row in sht_female_idle_area and half == 1:
                    continue
                if row in sht_male_idle_area and half == 2:
                    continue
                for cell in row:
                    if total == counts:
                        exit_flag = True
                        break
                    if cell.value == max:
                        cell.value = cell.value - 1
                        total = total - 1
                if exit_flag:
                    break
            max = max - 1
            for row in reversed(sht_area):
                if row in sht_female_idle_area and half == 1:
                    continue
                if row in sht_male_idle_area and half == 2:
                    continue
                for cell in reversed(row):
                    if total == counts:
                        exit_flag = True
                        break
                    if cell.value == max:
                        cell.value = cell.value - 1
                        total = total - 1
                if exit_flag:
                    break
            max = max - 1
            if exit_flag:
                break

    elif total < counts:
        min = start_pos.value
        for row in sht_area:
            if row in sht_idle_area:
                continue
            for cell in row:
                if min > cell.value:
                    min = cell.value

        exit_flag = False
        while total < counts:
            for row in sht_area:
                if row in sht_female_idle_area and half == 1:
                    continue
                if row in sht_male_idle_area and half == 2:
                    continue
                for cell in row:
                    if total == counts:
                        exit_flag = True
                        break
                    if cell.value == min:
                        cell.value = cell.value + 1
                        total = total + 1
                if exit_flag:
                    break;
            min = min + 1
            for row in reversed(sht_area):
                if row in sht_female_idle_area and half == 1:
                    continue
                if row in sht_male_idle_area and half == 2:
                    continue
                for cell in reversed(row):
                    if total == counts:
                        exit_flag = True
                        break
                    if cell.value == min:
                        cell.value = cell.value + 1
                        total = total + 1
                if exit_flag:
                    break
            min = min + 1
            if exit_flag:
                break


def adjust_sht_total(sht_total, list, hongyan_flag):
    dict_dtr_male = {'和平': 3, '沈河': 4, '大东': 5, '皇姑': 6, '铁西': 7, '浑南': 8, '于洪': 9, '苏家': 10, '沈北': 11}
    # dict_dtr_female = {'和平': 14, '沈河': 15, '大东': 16, '皇姑': 17, '铁西': 18, '浑南': 19, '于洪': 20, '苏家': 21, '沈北': 22}
    for key in dict_dtr_male:
        if list[0][0:2] == key:
            list_row = dict_dtr_male[key]
            if list[1] == '女':
                list_row += 11
            list_column = int(list[2] / 5)
            if (hongyan_flag == True):
                list_column += 1
            sht_total.cell(row=list_row, column=list_column).value = float(sht_total.cell(row=list_row, column=list_column).value) - 1
            break

    return sht_total


 # 根据表三调整表一
def allo_completed(sht_total, sht_cpl, hongyan_flag):
    notnull_max_row = sht_cpl.max_row
    exit_flag = False
    for row in sht_cpl.iter_rows(min_row=9, min_col=2, max_row=sht_cpl.max_row, max_col=2):
        for cell in row:
            if cell.value == "" or cell.value is None:
                notnull_max_row = cell.row - 1
                exit_flag = True
                break
        if exit_flag:
            break
    for row in sht_cpl.iter_rows(min_row=9, min_col=10, max_row=notnull_max_row, max_col=12):
        list = []
        for cell in row:
            if hongyan_flag == False:
                if sht_cpl.cell(row = cell.row,column = 27).value == "否" or sht_cpl.cell(row = cell.row,column = 27).value == "":
                    list.append(cell.value)
            else:
                if sht_cpl.cell(row=cell.row, column=27).value == "是":
                    list.append(cell.value)
        sht_total = adjust_sht_total(sht_total, list, hongyan_flag)
    return sht_total


root = tk.Tk()
root.withdraw()
print("导入第一个表")

file_path = filedialog.askopenfilename()
print("稍等几秒钟")
wb = load_workbook(filename=file_path)
sheet_origin = wb.active
sheet_origin.title = "总配额"


# 判断是否为鸿雁
hongyan_flag = False
hongyan_input = input('是否为鸿雁(如果有则输入小写y, 默认为否):')
if hongyan_input == "y":
    hongyan_flag = True

# 第三个表
sheet_cpl_flag = input('是否有已完成的表格(如果有则输入小写y, 默认为无):')
if sheet_cpl_flag == "y":
    file_path2 = filedialog.askopenfilename()
    wb_cpl = load_workbook(filename=file_path2)
    sheet_cpl = wb_cpl['被访者资料表']
    sheet_total = wb.copy_worksheet(sheet_origin)
    sheet_total.title = "表一三结合的配额"
    sheet_total = allo_completed(sheet_total, sheet_cpl, hongyan_flag)


# 新建两个sheet
sheet = wb.copy_worksheet(sheet_origin)
sheet.title = "单月配额"
sheet_even = wb.copy_worksheet(sheet_origin)
sheet_even.title = "双月配额"

if hongyan_flag == False:
    male_area = sheet['D3':'M11']
    male_area_even = sheet_even['D3':'M11']
    female_area = sheet['D14':'M22']
    female_area_even = sheet_even['D14':'M22']
else:
    male_area = sheet['E3':'N11']
    female_area = sheet['E14':'N22']
    male_area_even = sheet_even['E3':'N11']
    female_area_even = sheet_even['E14':'N22']

# 配额次数
times = 6
times = input('请输入配额次数(默认为数字6):')
if times == "":
    times = 6
else:
    times = int(times)

# 配额个数
counts = 172
counts = input('请输入预定的每月配额个数(默认为数字172):')
if counts == "":
    counts = 172
else:
    counts = int(counts)
sin_counts = 0
dou_counts = 0

# 单月
sin_counts = allocate(male_area, sin_counts, 0.3)
sin_counts = allocate(female_area, sin_counts, 0.5)
allo_counts(sheet, sin_counts, counts, 1, hongyan_flag)

# 双月
dou_counts = allocate(male_area_even, dou_counts, 0.5)
dou_counts = allocate(female_area_even, dou_counts, 0.3)
allo_counts(sheet_even, dou_counts, counts, 2, hongyan_flag)

filename = '配额结果.xlsx'
wb.save(filename=filename)
os.startfile(filename)
